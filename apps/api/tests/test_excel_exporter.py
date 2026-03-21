"""Tests pour excel_exporter.py — Export Excel multi-onglets.

Couvre:
- Génération du classeur avec données complètes
- Gestion des données partielles (onglets absents)
- Styles et mise en forme
- Graphiques openpyxl intégrés
- Robustesse face aux données manquantes/malformées
"""
import pytest
from io import BytesIO
from unittest.mock import MagicMock, patch
from dataclasses import dataclass, field
from datetime import datetime


# ── Fixtures ──────────────────────────────────────────────────────────────────

@dataclass
class MockProject:
    title: str = "Construction école Jean Jaurès"
    reference: str = "AO-2026-042"
    status: str = "ready"
    org_id: str = "00000000-0000-0000-0000-000000000001"


@dataclass
class MockChecklistItem:
    category: str = "Administratif"
    criticality: str = "éliminatoire"
    requirement: str = "Attestation URSSAF à jour"
    status: str = "OK"
    confidence: float = 92.0
    citation: str = "Article 4.2 du RC"


def _make_export_data(**overrides):
    """Build a mock ExportData-like object."""
    from app.services.export_data import ExportData

    defaults = dict(
        project=MockProject(),
        documents=[],
        summary={"project_overview": {"deadline_submission": "2026-06-15"}, "confidence_overall": 87.5},
        criteria={"evaluation": {"scoring_criteria": [
            {"name": "Prix", "weight": 40, "estimated_score": 75, "weighted_score": 30, "recommendation": "Optimiser le BPU"},
            {"name": "Valeur technique", "weight": 50, "estimated_score": 80, "weighted_score": 40, "recommendation": "Renforcer la méthodologie"},
            {"name": "Délai", "weight": 10, "estimated_score": 90, "weighted_score": 9, "recommendation": "Conforme"},
        ]}},
        gonogo={
            "score": 72,
            "decision": "GO conditionnel",
            "dimension_scores": {
                "Capacité financière": 80,
                "Certifications": 90,
                "Références similaires": 65,
                "Charge actuelle": 70,
                "Zone géographique": 75,
            },
        },
        timeline={"submission_deadline": "2026-06-15"},
        checklist_items=[MockChecklistItem(), MockChecklistItem(
            category="Technique", criticality="important",
            requirement="BIM niveau 2", status="MANQUANT", confidence=45.0,
            citation="CCTP art. 3.4",
        )],
        checklist_stats={"eliminatoire": 1, "important": 1, "info": 0, "ok": 1},
        confidence=87.5,
        ccap_analysis=None,
        ccag_derogations=None,
        ccap_clauses_risquees=[
            {"article": "Art. 10.2", "clause": "Pénalités de retard 1/500e", "severity": "élevé",
             "impact": "Risque financier", "recommendation": "Négocier le taux"},
        ],
        rc_analysis=None,
        ae_analysis=None,
        cctp_analysis=None,
        dc_check=None,
        conflicts={"conflicts": [
            {"doc_source": "CCAP", "doc_target": "CCTP", "conflict_type": "délais",
             "severity": "high", "description": "Incohérence sur le délai global"},
        ]},
        cashflow={
            "monthly_cashflow": [
                {"month": 1, "income_eur": 0, "expenses_eur": 50000, "balance_eur": -50000, "cumulative_eur": -50000},
                {"month": 2, "income_eur": 80000, "expenses_eur": 45000, "balance_eur": 35000, "cumulative_eur": -15000},
                {"month": 3, "income_eur": 90000, "expenses_eur": 40000, "balance_eur": 50000, "cumulative_eur": 35000},
                {"month": 4, "income_eur": 85000, "expenses_eur": 42000, "balance_eur": 43000, "cumulative_eur": 78000},
            ]
        },
        subcontracting=None,
        questions=None,
        scoring={"criteria": [
            {"name": "Prix", "weight": 40, "estimated_score": 75, "weighted_score": 30},
            {"name": "Technique", "weight": 50, "estimated_score": 80, "weighted_score": 40},
        ]},
        dpgf_pricing=[
            {"designation": "Gros-oeuvre", "prix_saisi": 120, "prix_min": 100, "prix_max": 150,
             "status": "OK", "ecart_pct": 0},
            {"designation": "VRD", "prix_saisi": 50, "prix_min": 80, "prix_max": 120,
             "status": "SOUS_EVALUE", "ecart_pct": -37.5},
        ],
        glossaire_btp=None,
        days_remaining=42,
        deadline_str="2026-06-15",
        gonogo_obj=None,
        timeline_obj=None,
    )
    defaults.update(overrides)
    return ExportData(**defaults)


# ── Tests ─────────────────────────────────────────────────────────────────────

class TestGenerateAnalysisExcel:
    """Tests pour generate_analysis_excel()."""

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_generates_valid_xlsx(self, mock_fetch):
        """Le fichier généré est un XLSX valide lisible par openpyxl."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")

        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) >= 1

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_all_sheets_present_with_full_data(self, mock_fetch):
        """Avec des données complètes, les 6 onglets sont créés."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        expected_sheets = {"Synthèse", "Checklist", "Scoring", "Trésorerie", "DPGF Pricing", "Risques"}
        assert set(wb.sheetnames) == expected_sheets

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_minimal_data_only_synthese(self, mock_fetch):
        """Avec des données minimales, seul l'onglet Synthèse est présent."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data(
            checklist_items=[], scoring=None, cashflow=None,
            dpgf_pricing=None, ccap_clauses_risquees=None, conflicts=None,
        )
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        assert "Synthèse" in wb.sheetnames
        assert "Checklist" not in wb.sheetnames

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_synthese_contains_project_title(self, mock_fetch):
        """L'onglet Synthèse contient le titre du projet."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Synthèse"]
        assert "Construction école Jean Jaurès" in str(ws["A1"].value)

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_gonogo_score_in_synthese(self, mock_fetch):
        """Le score Go/No-Go apparaît dans l'onglet Synthèse."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Synthèse"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(1, 30) if ws.cell(row=r, column=2).value]
        assert any("72/100" in v for v in values)

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_checklist_has_autofilter(self, mock_fetch):
        """L'onglet Checklist a un filtre automatique activé."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Checklist"]
        assert ws.auto_filter.ref is not None

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_cashflow_has_chart(self, mock_fetch):
        """L'onglet Trésorerie contient un graphique LineChart."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Trésorerie"]
        assert len(ws._charts) >= 1

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_scoring_has_chart(self, mock_fetch):
        """L'onglet Scoring contient un graphique BarChart."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Scoring"]
        assert len(ws._charts) >= 1

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_dpgf_pricing_colors(self, mock_fetch):
        """Les cellules DPGF ont des couleurs conditionnelles selon le statut."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["DPGF Pricing"]
        # Row 4 = first data row, col 5 = status
        status_fill = ws.cell(row=4, column=5).fill
        assert status_fill.fgColor is not None

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_risques_sheet_has_ccap_and_conflicts(self, mock_fetch):
        """L'onglet Risques contient les clauses CCAP et les conflits."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data()
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))

        ws = wb["Risques"]
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v])

        assert any("Art. 10.2" in v for v in all_values)
        assert any("CCAP" in v for v in all_values)


class TestExcelRobustness:
    """Tests de robustesse — données manquantes ou malformées."""

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_none_dpgf_entries_skipped(self, mock_fetch):
        """Les entrées DPGF non-dict sont ignorées sans crash."""
        from app.services.excel_exporter import generate_analysis_excel

        mock_fetch.return_value = _make_export_data(
            dpgf_pricing=[None, "invalid", {"designation": "Valid", "status": "OK"}]
        )
        result = generate_analysis_excel(MagicMock(), "test-id")
        assert isinstance(result, bytes)
        assert len(result) > 0

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_empty_cashflow_no_sheet(self, mock_fetch):
        """Pas d'onglet Trésorerie si cashflow vide."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data(cashflow={"monthly_cashflow": []})
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))
        assert "Trésorerie" not in wb.sheetnames

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_no_gonogo_still_works(self, mock_fetch):
        """Sans Go/No-Go, l'onglet Synthèse est toujours généré."""
        from app.services.excel_exporter import generate_analysis_excel

        mock_fetch.return_value = _make_export_data(gonogo=None)
        result = generate_analysis_excel(MagicMock(), "test-id")
        assert isinstance(result, bytes)
        assert len(result) > 0

    @patch("app.services.excel_exporter.fetch_export_data")
    def test_single_month_cashflow(self, mock_fetch):
        """Un seul mois de cashflow génère les données sans graphique."""
        from app.services.excel_exporter import generate_analysis_excel
        from openpyxl import load_workbook

        mock_fetch.return_value = _make_export_data(cashflow={
            "monthly_cashflow": [{"month": 1, "cumulative_eur": -20000}]
        })
        result = generate_analysis_excel(MagicMock(), "test-id")
        wb = load_workbook(BytesIO(result))
        assert "Trésorerie" in wb.sheetnames
        # Un seul mois → pas de graphique (min 2 requis)
        ws = wb["Trésorerie"]
        assert len(ws._charts) == 0
