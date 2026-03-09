"""Extraction et export Excel des tableaux DPGF / BPU depuis un PDF.

Utilise pdfplumber pour détecter et extraire les tableaux tabulaires,
normalise les colonnes (différents acheteurs utilisent des en-têtes
variés) et génère un fichier Excel professionnel avec openpyxl.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import pdfplumber
import structlog

logger = structlog.get_logger(__name__)

# ── Constantes de mise en forme ──────────────────────────────────────────────

HEADER_FILL_HEX = "1E40AF"   # Bleu primaire — cohérent avec le reste de l'app
ALT_ROW_FILL_HEX = "F1F5F9"  # Gris très clair pour les lignes paires
TOTAL_FILL_HEX = "DBEAFE"    # Bleu pâle pour la ligne total HT

# ── Mapping des variantes de noms de colonnes ────────────────────────────────

# Chaque entrée : regex (case-insensitive) → clé canonique interne
COLUMN_ALIASES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"num[ée]ro|n[°o]\.?\s*lot|r[ée]f[ée]rence|article", re.IGNORECASE), "numero"),
    (re.compile(r"d[ée]signation|libell[ée]|description|intitul[ée]|prestation", re.IGNORECASE), "designation"),
    (re.compile(r"unit[ée]|u\.?", re.IGNORECASE), "unite"),
    (re.compile(r"quantit[ée]|qte|qté|qty", re.IGNORECASE), "quantite"),
    (re.compile(r"prix\s*unit(?:aire)?|pu\b|p\.u\.?", re.IGNORECASE), "prix_unitaire"),
    (re.compile(r"montant\s*h\.?t\.?|total\s*h\.?t\.?|prix\s*total|sous.total", re.IGNORECASE), "montant_ht"),
    (re.compile(r"prix\b(?!\s*unit)", re.IGNORECASE), "prix_unitaire"),  # BPU : colonne "Prix"
]

# En-têtes canoniques affichés dans l'Excel (DPGF)
DPGF_HEADERS = ["N°", "Désignation des travaux", "Unité", "Quantité", "Prix unitaire HT", "Montant HT"]
DPGF_KEYS = ["numero", "designation", "unite", "quantite", "prix_unitaire", "montant_ht"]

# En-têtes canoniques affichés dans l'Excel (BPU)
BPU_HEADERS = ["N°", "Désignation", "Unité", "Prix unitaire HT"]
BPU_KEYS = ["numero", "designation", "unite", "prix_unitaire"]


# ── Structures de données ────────────────────────────────────────────────────

@dataclass
class DpgfRow:
    numero: str = ""
    designation: str = ""
    unite: str = ""
    quantite: str = ""
    prix_unitaire: str = ""
    montant_ht: str = ""


@dataclass
class BpuRow:
    numero: str = ""
    designation: str = ""
    unite: str = ""
    prix_unitaire: str = ""


@dataclass
class ExtractedTable:
    doc_type: str               # "DPGF" | "BPU"
    rows: list[DpgfRow | BpuRow] = field(default_factory=list)
    source_page: int = 0
    raw_headers: list[str] = field(default_factory=list)


# ── Parsing / normalisation ──────────────────────────────────────────────────

def _normalize_col_name(header: str) -> str | None:
    """Tente de mapper un en-tête brut vers une clé canonique."""
    header = (header or "").strip()
    for pattern, canonical in COLUMN_ALIASES:
        if pattern.search(header):
            return canonical
    return None


def _parse_number(value: str) -> float | None:
    """Convertit une chaîne de type '1 234,56 €' en float Python."""
    if not value:
        return None
    cleaned = re.sub(r"[€$\s]", "", str(value))
    cleaned = cleaned.replace("\xa0", "").replace(" ", "")
    cleaned = cleaned.replace(",", ".")
    # Gère le format 1.234.567,89 → supprime tous les points sauf le dernier
    parts = cleaned.split(".")
    if len(parts) > 2:
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _is_header_row(row: list[str | None]) -> bool:
    """Détecte si une ligne de tableau correspond à une ligne d'en-têtes."""
    non_empty = [c for c in row if c and c.strip()]
    if not non_empty:
        return False
    matched = sum(1 for c in non_empty if _normalize_col_name(c) is not None)
    # On considère un header si ≥ 2 colonnes sont reconnues
    return matched >= 2


def _guess_doc_type_from_headers(headers: list[str]) -> str:
    """Devine DPGF ou BPU selon les colonnes présentes."""
    keys = [_normalize_col_name(h) for h in headers]
    # Présence de Quantité → DPGF, sinon BPU
    if "quantite" in keys or "montant_ht" in keys:
        return "DPGF"
    return "BPU"


def _map_row_to_dataclass(
    raw_row: list[str | None],
    col_mapping: dict[int, str],
    doc_type: str,
) -> DpgfRow | BpuRow | None:
    """Crée un DpgfRow ou BpuRow depuis une ligne brute + mapping d'indices."""
    values: dict[str, str] = {}
    for idx, key in col_mapping.items():
        if idx < len(raw_row):
            values[key] = (raw_row[idx] or "").strip()

    # Ignorer les lignes complètement vides
    if not any(values.values()):
        return None

    if doc_type == "DPGF":
        return DpgfRow(
            numero=values.get("numero", ""),
            designation=values.get("designation", ""),
            unite=values.get("unite", ""),
            quantite=values.get("quantite", ""),
            prix_unitaire=values.get("prix_unitaire", ""),
            montant_ht=values.get("montant_ht", ""),
        )
    else:  # BPU
        return BpuRow(
            numero=values.get("numero", ""),
            designation=values.get("designation", ""),
            unite=values.get("unite", ""),
            prix_unitaire=values.get("prix_unitaire", ""),
        )


# ── Extraction PDF ───────────────────────────────────────────────────────────

def extract_tables_from_pdf(pdf_bytes: bytes, filename: str = "document.pdf") -> list[ExtractedTable]:
    """
    Parcourt toutes les pages du PDF et extrait les tableaux DPGF/BPU.

    Retourne une liste de ExtractedTable (peut être vide si aucun tableau
    pertinent n'est trouvé).
    """
    extracted: list[ExtractedTable] = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if not tables:
                    continue

                for raw_table in tables:
                    if not raw_table or len(raw_table) < 2:
                        continue

                    # Chercher la ligne d'en-tête dans les 3 premières lignes
                    header_idx: int | None = None
                    for i, row in enumerate(raw_table[:3]):
                        if _is_header_row(row):
                            header_idx = i
                            break

                    if header_idx is None:
                        logger.debug(
                            "dpgf_extractor.skip_table_no_header",
                            page=page_num,
                            filename=filename,
                        )
                        continue

                    raw_headers = [str(c or "").strip() for c in raw_table[header_idx]]
                    doc_type = _guess_doc_type_from_headers(raw_headers)

                    # Construire le mapping indice→clé canonique
                    col_mapping: dict[int, str] = {}
                    for i, h in enumerate(raw_headers):
                        key = _normalize_col_name(h)
                        if key and key not in col_mapping.values():
                            col_mapping[i] = key

                    if not col_mapping:
                        continue

                    rows: list[DpgfRow | BpuRow] = []
                    for raw_row in raw_table[header_idx + 1:]:
                        obj = _map_row_to_dataclass(raw_row, col_mapping, doc_type)
                        if obj is not None:
                            rows.append(obj)

                    if rows:
                        extracted.append(ExtractedTable(
                            doc_type=doc_type,
                            rows=rows,
                            source_page=page_num,
                            raw_headers=raw_headers,
                        ))

    except Exception as exc:
        logger.error(
            "dpgf_extractor.pdf_parse_error",
            filename=filename,
            error=str(exc),
        )
        raise ValueError(f"Impossible d'analyser le PDF pour extraction DPGF/BPU : {exc}") from exc

    logger.info(
        "dpgf_extractor.extraction_done",
        filename=filename,
        table_count=len(extracted),
        total_rows=sum(len(t.rows) for t in extracted),
    )
    return extracted


# ── Génération Excel ─────────────────────────────────────────────────────────

def _hex_to_argb(hex_color: str) -> str:
    """Convertit 'RRGGBB' → 'FFRRGGBB' attendu par openpyxl."""
    return "FF" + hex_color.upper()


def generate_excel(tables: list[ExtractedTable], project_title: str = "") -> bytes:
    """
    Génère un fichier Excel (.xlsx) à partir de la liste de tableaux extraits.

    Structure :
    - Une feuille par tableau détecté ("DPGF", "DPGF_2", "BPU", …)
    - En-tête coloré (#1E40AF), lignes alternées gris clair
    - Colonnes numériques formatées
    - Pied de page avec total HT calculé (DPGF uniquement)

    Retourne les bytes du fichier .xlsx.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl n'est pas installé. Ajoutez openpyxl>=3.1.0 à requirements.txt."
        ) from exc

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]  # Supprimer la feuille vide par défaut

    # Styles partagés
    header_font = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    header_fill = PatternFill("solid", fgColor=_hex_to_argb(HEADER_FILL_HEX))
    alt_fill = PatternFill("solid", fgColor=_hex_to_argb(ALT_ROW_FILL_HEX))
    total_fill = PatternFill("solid", fgColor=_hex_to_argb(TOTAL_FILL_HEX))
    total_font = Font(bold=True, size=10, name="Calibri")
    data_font = Font(size=10, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border_side = Side(style="thin", color="FFD1D5DB")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    num_format_currency = '#,##0.00 "€"'
    num_format_qty = '#,##0.##'

    # Compteur pour déduplication des noms d'onglets
    sheet_name_counts: dict[str, int] = {}

    for table in tables:
        base_name = table.doc_type
        count = sheet_name_counts.get(base_name, 0)
        sheet_name_counts[base_name] = count + 1
        ws_name = base_name if count == 0 else f"{base_name}_{count + 1}"

        ws = wb.create_sheet(title=ws_name)

        # ── Ligne de titre du projet ────────────────────────────────────
        if project_title:
            ws.merge_cells("A1:F1")
            title_cell = ws["A1"]
            title_cell.value = f"AO Copilot — {table.doc_type}  |  {project_title}"
            title_cell.font = Font(bold=True, size=12, name="Calibri", color=_hex_to_argb(HEADER_FILL_HEX))
            title_cell.alignment = left_align
            ws.row_dimensions[1].height = 22
            title_row_offset = 1
        else:
            title_row_offset = 0

        # ── Ligne d'information d'extraction ───────────────────────────
        meta_row = title_row_offset + 1
        ws.cell(row=meta_row, column=1).value = (
            f"Extrait depuis la page {table.source_page}  —  "
            f"Généré par AO Copilot le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        ws.cell(row=meta_row, column=1).font = Font(italic=True, size=9, name="Calibri", color="FF6B7280")
        ws.row_dimensions[meta_row].height = 16

        header_row = meta_row + 1

        is_dpgf = table.doc_type == "DPGF"
        col_keys = DPGF_KEYS if is_dpgf else BPU_KEYS
        col_headers = DPGF_HEADERS if is_dpgf else BPU_HEADERS
        n_cols = len(col_headers)

        # ── En-têtes ───────────────────────────────────────────────────
        for col_idx, header_text in enumerate(col_headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 20

        # ── Données ────────────────────────────────────────────────────
        total_ht: float = 0.0
        data_start_row = header_row + 1

        for row_idx, row_obj in enumerate(table.rows):
            excel_row = data_start_row + row_idx
            use_alt = row_idx % 2 == 1

            row_data: list[Any] = [getattr(row_obj, k, "") for k in col_keys]

            for col_idx, raw_value in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col_idx)
                key = col_keys[col_idx - 1]

                # Essayer de convertir en nombre pour les colonnes numériques
                if key in ("quantite", "prix_unitaire", "montant_ht"):
                    num = _parse_number(str(raw_value))
                    if num is not None:
                        cell.value = num
                        if key == "quantite":
                            cell.number_format = num_format_qty
                        else:
                            cell.number_format = num_format_currency
                        cell.alignment = right_align
                        if key == "montant_ht":
                            total_ht += num
                    else:
                        cell.value = str(raw_value)
                        cell.alignment = right_align
                elif key == "numero":
                    cell.value = str(raw_value)
                    cell.alignment = center_align
                else:
                    cell.value = str(raw_value)
                    cell.alignment = left_align

                cell.font = data_font
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill

            ws.row_dimensions[excel_row].height = 18

        # ── Ligne TOTAL HT (DPGF seulement) ───────────────────────────
        if is_dpgf and table.rows:
            total_row = data_start_row + len(table.rows)
            # Fusionner les colonnes N° → Prix unitaire
            merge_end_col = n_cols - 1  # avant-dernière colonne
            if merge_end_col >= 1:
                ws.merge_cells(
                    start_row=total_row, start_column=1,
                    end_row=total_row, end_column=merge_end_col,
                )
            label_cell = ws.cell(row=total_row, column=1)
            label_cell.value = "TOTAL HT"
            label_cell.font = total_font
            label_cell.fill = total_fill
            label_cell.alignment = right_align
            label_cell.border = thin_border

            total_cell = ws.cell(row=total_row, column=n_cols)
            total_cell.value = total_ht
            total_cell.number_format = num_format_currency
            total_cell.font = total_font
            total_cell.fill = total_fill
            total_cell.alignment = right_align
            total_cell.border = thin_border

            ws.row_dimensions[total_row].height = 20

        # ── Largeurs de colonnes ───────────────────────────────────────
        col_widths: dict[str, int] = {
            "numero": 8,
            "designation": 45,
            "unite": 8,
            "quantite": 12,
            "prix_unitaire": 18,
            "montant_ht": 18,
        }
        for col_idx, key in enumerate(col_keys, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 15)

        # ── Figer les volets (en-têtes visibles au scroll) ────────────
        freeze_cell = ws.cell(row=header_row + 1, column=1)
        ws.freeze_panes = freeze_cell

        # ── Filtre automatique ─────────────────────────────────────────
        last_data_row = data_start_row + len(table.rows) - 1
        if last_data_row >= data_start_row:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(n_cols)}{last_data_row}"
            )

    # Si aucun tableau n'a été extrait, créer une feuille d'information
    if not tables:
        ws = wb.create_sheet(title="Aucun tableau détecté")
        ws["A1"].value = (
            "Aucun tableau DPGF ou BPU n'a pu être extrait automatiquement depuis ce PDF."
        )
        ws["A1"].font = Font(bold=True, size=11, name="Calibri", color="FFDC2626")
        ws["A2"].value = (
            "Vérifiez que le document contient bien des tableaux structurés "
            "(et non des images de tableaux)."
        )
        ws["A2"].font = Font(size=10, name="Calibri", color="FF6B7280")
        ws.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Point d'entrée principal ─────────────────────────────────────────────────

def extract_dpgf(
    pdf_bytes: bytes,
    filename: str = "document.pdf",
    project_title: str = "",
) -> bytes:
    """
    Extrait les tableaux DPGF/BPU du PDF et retourne les bytes d'un fichier Excel.

    Arguments :
        pdf_bytes     : contenu brut du fichier PDF
        filename      : nom original du fichier (pour les logs)
        project_title : titre du projet AO (affiché en en-tête Excel)

    Retourne :
        bytes du fichier .xlsx
    """
    tables = extract_tables_from_pdf(pdf_bytes, filename)
    return generate_excel(tables, project_title=project_title)
