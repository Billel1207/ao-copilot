"""Export Excel enrichi — scoring, cashflow, checklist, Go/No-Go, DPGF pricing.

Génère un classeur Excel multi-onglets avec mise en forme professionnelle et
graphiques openpyxl intégrés. Complète l'export DPGF brut (dpgf_extractor.py)
avec les résultats d'analyse IA.

Usage:
    from app.services.excel_exporter import generate_analysis_excel
    xlsx_bytes = generate_analysis_excel(db, project_id)
"""
from __future__ import annotations

import structlog
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.report_theme import get_theme
from app.services.export_data import fetch_export_data, ExportData

logger = structlog.get_logger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────

_theme = get_theme()
_HEADER_FILL = PatternFill("solid", fgColor=_theme.header_bg.lstrip("#"))
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=_theme.header_bg.lstrip("#"))
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color=_theme.primary.lstrip("#"))
_BODY_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", bold=True, size=10)
_ALT_FILL = PatternFill("solid", fgColor="F1F5F9")
_GO_FILL = PatternFill("solid", fgColor=_theme.go_bg.lstrip("#"))
_NOGO_FILL = PatternFill("solid", fgColor=_theme.nogo_bg.lstrip("#"))
_WARN_FILL = PatternFill("solid", fgColor=_theme.risk_med_bg.lstrip("#"))
_DANGER_FILL = PatternFill("solid", fgColor=_theme.risk_high_bg.lstrip("#"))
_SUCCESS_FILL = PatternFill("solid", fgColor=_theme.risk_low_bg.lstrip("#"))
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_data_row(ws, row: int, col_count: int, alt: bool = False):
    """Apply data row styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = _ALT_FILL


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ── Sheets ────────────────────────────────────────────────────────────────────

def _add_synthese_sheet(wb: Workbook, data: ExportData):
    """Onglet Synthèse — résumé projet, Go/No-Go, métriques clés."""
    ws = wb.active
    ws.title = "Synthèse"

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Rapport d'analyse — {data.project.title}"
    ws["A1"].font = _TITLE_FONT

    ws["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %Hh%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="94A3B8")

    row = 4

    # Infos projet
    infos = [
        ("Projet", data.project.title),
        ("Référence", data.project.reference or "—"),
        ("Statut", data.project.status),
    ]
    if data.deadline_str:
        infos.append(("Date limite", data.deadline_str))
    if data.days_remaining is not None:
        infos.append(("Jours restants", data.days_remaining))
    if data.confidence:
        infos.append(("Confiance IA", f"{data.confidence}%"))

    for label, value in infos:
        ws.cell(row=row, column=1, value=label).font = _BOLD_FONT
        ws.cell(row=row, column=2, value=str(value)).font = _BODY_FONT
        row += 1

    # Go/No-Go
    if data.gonogo:
        row += 1
        ws.cell(row=row, column=1, value="Score Go/No-Go").font = _SUBTITLE_FONT
        row += 1

        score = data.gonogo.get("score", 0)
        decision = data.gonogo.get("decision", "—")
        ws.cell(row=row, column=1, value="Score global").font = _BOLD_FONT
        score_cell = ws.cell(row=row, column=2, value=f"{score}/100")
        score_cell.font = _BOLD_FONT
        if score >= 70:
            score_cell.fill = _GO_FILL
        elif score >= 50:
            score_cell.fill = _WARN_FILL
        else:
            score_cell.fill = _NOGO_FILL
        row += 1

        ws.cell(row=row, column=1, value="Décision").font = _BOLD_FONT
        ws.cell(row=row, column=2, value=str(decision).upper()).font = _BOLD_FONT
        row += 1

        # Dimensions breakdown
        dims = data.gonogo.get("dimension_scores") or data.gonogo.get("breakdown") or {}
        if isinstance(dims, dict) and dims:
            row += 1
            headers = ["Dimension", "Score"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c, value=h)
            _style_header_row(ws, row, len(headers))
            row += 1

            for i, (dim, val) in enumerate(dims.items()):
                ws.cell(row=row, column=1, value=str(dim)).font = _BODY_FONT
                val_cell = ws.cell(row=row, column=2, value=float(val) if val else 0)
                val_cell.font = _BODY_FONT
                val_cell.alignment = _CENTER
                _style_data_row(ws, row, 2, alt=i % 2 == 1)
                row += 1

    # Checklist stats
    row += 1
    ws.cell(row=row, column=1, value="Checklist — Statistiques").font = _SUBTITLE_FONT
    row += 1
    for label, count in data.checklist_stats.items():
        ws.cell(row=row, column=1, value=label.capitalize()).font = _BODY_FONT
        ws.cell(row=row, column=2, value=count).font = _BODY_FONT
        ws.cell(row=row, column=2).alignment = _CENTER
        row += 1

    _auto_width(ws)


def _add_checklist_sheet(wb: Workbook, data: ExportData):
    """Onglet Checklist — tous les items avec criticité et statut."""
    if not data.checklist_items:
        return

    ws = wb.create_sheet("Checklist")
    headers = ["Catégorie", "Criticité", "Exigence", "Statut", "Confiance", "Citation"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, item in enumerate(data.checklist_items, start=2):
        ws.cell(row=i, column=1, value=item.category or "—").font = _BODY_FONT
        crit_cell = ws.cell(row=i, column=2, value=item.criticality or "—")
        crit_cell.font = _BODY_FONT
        crit = (item.criticality or "").lower()
        if "liminatoire" in crit:
            crit_cell.fill = _DANGER_FILL
        elif "important" in crit:
            crit_cell.fill = _WARN_FILL

        ws.cell(row=i, column=3, value=item.requirement or "—").font = _BODY_FONT
        ws.cell(row=i, column=3).alignment = _WRAP

        status_cell = ws.cell(row=i, column=4, value=item.status or "—")
        status_cell.font = _BODY_FONT
        status_cell.alignment = _CENTER
        if (item.status or "").upper() == "OK":
            status_cell.fill = _SUCCESS_FILL

        ws.cell(row=i, column=5, value=f"{item.confidence}%" if item.confidence else "—").font = _BODY_FONT
        ws.cell(row=i, column=6, value=item.citation or "—").font = _BODY_FONT
        ws.cell(row=i, column=6).alignment = _WRAP

        _style_data_row(ws, i, len(headers), alt=i % 2 == 0)

    _auto_width(ws)
    ws.auto_filter.ref = f"A1:F{len(data.checklist_items) + 1}"


def _add_scoring_sheet(wb: Workbook, data: ExportData):
    """Onglet Scoring — simulation de notation avec graphique."""
    if not data.scoring:
        return

    ws = wb.create_sheet("Scoring")
    ws.cell(row=1, column=1, value="Simulation de notation").font = _TITLE_FONT

    criteria_list = data.scoring.get("criteria") or data.scoring.get("scoring_criteria") or []
    if not criteria_list:
        return

    headers = ["Critère", "Pondération (%)", "Note estimée", "Note pondérée", "Recommandation"]
    row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(headers))

    for i, crit in enumerate(criteria_list):
        r = row + 1 + i
        if isinstance(crit, dict):
            ws.cell(row=r, column=1, value=crit.get("name", crit.get("criterion", "—"))).font = _BODY_FONT
            ws.cell(row=r, column=2, value=crit.get("weight", crit.get("ponderation", 0))).font = _BODY_FONT
            ws.cell(row=r, column=2).alignment = _CENTER
            ws.cell(row=r, column=3, value=crit.get("estimated_score", crit.get("note", 0))).font = _BODY_FONT
            ws.cell(row=r, column=3).alignment = _CENTER
            weighted = crit.get("weighted_score", crit.get("note_ponderee", 0))
            ws.cell(row=r, column=4, value=weighted).font = _BOLD_FONT
            ws.cell(row=r, column=4).alignment = _CENTER
            ws.cell(row=r, column=5, value=crit.get("recommendation", crit.get("conseil", ""))).font = _BODY_FONT
            ws.cell(row=r, column=5).alignment = _WRAP
        _style_data_row(ws, r, len(headers), alt=i % 2 == 1)

    # Bar chart for scoring
    data_end_row = row + len(criteria_list)
    if len(criteria_list) >= 2:
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Notes pondérées par critère"
            chart.y_axis.title = "Note pondérée"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=4, min_row=row, max_row=data_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=data_end_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.solidFill = _theme.primary.lstrip("#")
            ws.add_chart(chart, f"A{data_end_row + 3}")
        except Exception as e:
            logger.warning("excel_scoring_chart_error", error=str(e))

    _auto_width(ws)


def _add_cashflow_sheet(wb: Workbook, data: ExportData):
    """Onglet Trésorerie — simulation cashflow mensuel avec graphique."""
    if not data.cashflow:
        return

    monthly = (
        data.cashflow.get("monthly_cashflow")
        or data.cashflow.get("simulation", {}).get("monthly_cashflow")
        or data.cashflow.get("cashflow_mensuel")
        or []
    )
    if not monthly:
        return

    ws = wb.create_sheet("Trésorerie")
    ws.cell(row=1, column=1, value="Simulation trésorerie prévisionnelle").font = _TITLE_FONT

    headers = ["Mois", "Entrées (€)", "Sorties (€)", "Solde mensuel (€)", "Cumul (€)"]
    row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(headers))

    for i, m in enumerate(monthly):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=m.get("month", m.get("mois", i + 1))).font = _BODY_FONT
        ws.cell(row=r, column=1).alignment = _CENTER
        ws.cell(row=r, column=2, value=m.get("income_eur", m.get("entrees", 0))).font = _BODY_FONT
        ws.cell(row=r, column=2).number_format = '#,##0 €'
        ws.cell(row=r, column=3, value=m.get("expenses_eur", m.get("sorties", 0))).font = _BODY_FONT
        ws.cell(row=r, column=3).number_format = '#,##0 €'
        ws.cell(row=r, column=4, value=m.get("balance_eur", m.get("solde", 0))).font = _BODY_FONT
        ws.cell(row=r, column=4).number_format = '#,##0 €'
        cumul_cell = ws.cell(row=r, column=5, value=m.get("cumulative_eur", m.get("tresorerie_cumulee", 0)))
        cumul_cell.font = _BOLD_FONT
        cumul_cell.number_format = '#,##0 €'
        cumul_val = cumul_cell.value or 0
        if isinstance(cumul_val, (int, float)) and cumul_val < 0:
            cumul_cell.fill = _DANGER_FILL
        _style_data_row(ws, r, len(headers), alt=i % 2 == 1)

    # Line chart
    data_end_row = row + len(monthly)
    if len(monthly) >= 2:
        try:
            chart = LineChart()
            chart.title = "Trésorerie cumulée"
            chart.y_axis.title = "€"
            chart.style = 10
            chart.width = 22
            chart.height = 12

            cumul_ref = Reference(ws, min_col=5, min_row=row, max_row=data_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=data_end_row)
            chart.add_data(cumul_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            s = chart.series[0]
            s.graphicalProperties.line.solidFill = _theme.primary.lstrip("#")
            s.graphicalProperties.line.width = 25000  # EMU
            ws.add_chart(chart, f"A{data_end_row + 3}")
        except Exception as e:
            logger.warning("excel_cashflow_chart_error", error=str(e))

    _auto_width(ws)


def _add_dpgf_pricing_sheet(wb: Workbook, data: ExportData):
    """Onglet DPGF Pricing — benchmark prix vs référentiel BTP."""
    if not data.dpgf_pricing:
        return

    ws = wb.create_sheet("DPGF Pricing")
    ws.cell(row=1, column=1, value="Benchmark DPGF — Prix vs Référentiel BTP").font = _TITLE_FONT

    headers = ["Désignation", "Prix saisi (€)", "Prix min (€)", "Prix max (€)", "Statut", "Écart (%)"]
    row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(headers))

    for i, line in enumerate(data.dpgf_pricing):
        if not isinstance(line, dict):
            continue
        r = row + 1 + i
        ws.cell(row=r, column=1, value=line.get("designation", "—")).font = _BODY_FONT
        ws.cell(row=r, column=1).alignment = _WRAP

        prix = line.get("prix_saisi", line.get("prix_unitaire", 0))
        ws.cell(row=r, column=2, value=prix).number_format = '#,##0.00 €'
        ws.cell(row=r, column=3, value=line.get("prix_min", line.get("min", 0))).number_format = '#,##0.00 €'
        ws.cell(row=r, column=4, value=line.get("prix_max", line.get("max", 0))).number_format = '#,##0.00 €'

        status = line.get("status", "OK")
        status_cell = ws.cell(row=r, column=5, value=status)
        status_cell.font = _BOLD_FONT
        status_cell.alignment = _CENTER
        if status == "SOUS_EVALUE":
            status_cell.fill = _DANGER_FILL
        elif status == "SUR_EVALUE":
            status_cell.fill = _WARN_FILL
        else:
            status_cell.fill = _SUCCESS_FILL

        ecart = line.get("ecart_pct", line.get("deviation_pct"))
        ws.cell(row=r, column=6, value=f"{ecart}%" if ecart else "—").font = _BODY_FONT
        ws.cell(row=r, column=6).alignment = _CENTER

        _style_data_row(ws, r, len(headers), alt=i % 2 == 1)

    _auto_width(ws)
    count = sum(1 for l in data.dpgf_pricing if isinstance(l, dict))
    ws.auto_filter.ref = f"A{row}:F{row + count}"


def _add_risks_sheet(wb: Workbook, data: ExportData):
    """Onglet Risques — CCAP clauses risquées + conflits."""
    has_ccap = data.ccap_clauses_risquees and isinstance(data.ccap_clauses_risquees, list)
    has_conflicts = data.conflicts and isinstance(data.conflicts, dict)

    if not has_ccap and not has_conflicts:
        return

    ws = wb.create_sheet("Risques")
    row = 1

    if has_ccap:
        ws.cell(row=row, column=1, value="Clauses CCAP à risque").font = _TITLE_FONT
        row += 1

        headers = ["Article", "Clause", "Niveau", "Impact", "Recommandation"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, len(headers))
        row += 1

        for i, clause in enumerate(data.ccap_clauses_risquees):
            if not isinstance(clause, dict):
                continue
            ws.cell(row=row, column=1, value=clause.get("article", "—")).font = _BODY_FONT
            ws.cell(row=row, column=2, value=clause.get("clause", clause.get("titre", "—"))).font = _BODY_FONT
            ws.cell(row=row, column=2).alignment = _WRAP

            sev = clause.get("severity", clause.get("niveau", "—"))
            sev_cell = ws.cell(row=row, column=3, value=str(sev))
            sev_cell.font = _BOLD_FONT
            sev_cell.alignment = _CENTER
            sev_lower = str(sev).lower()
            if "critique" in sev_lower or "high" in sev_lower or "élevé" in sev_lower:
                sev_cell.fill = _DANGER_FILL
            elif "moyen" in sev_lower or "medium" in sev_lower:
                sev_cell.fill = _WARN_FILL

            ws.cell(row=row, column=4, value=clause.get("impact", "—")).font = _BODY_FONT
            ws.cell(row=row, column=4).alignment = _WRAP
            ws.cell(row=row, column=5, value=clause.get("recommendation", clause.get("recommandation", "—"))).font = _BODY_FONT
            ws.cell(row=row, column=5).alignment = _WRAP

            _style_data_row(ws, row, len(headers), alt=i % 2 == 1)
            row += 1

    if has_conflicts:
        row += 1
        ws.cell(row=row, column=1, value="Conflits inter-documents").font = _TITLE_FONT
        row += 1

        conflict_list = data.conflicts.get("conflicts") or data.conflicts.get("items") or []
        if conflict_list:
            headers = ["Document source", "Document cible", "Type", "Sévérité", "Description"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c, value=h)
            _style_header_row(ws, row, len(headers))
            row += 1

            for i, conflict in enumerate(conflict_list):
                if not isinstance(conflict, dict):
                    continue
                ws.cell(row=row, column=1, value=conflict.get("doc_source", "—")).font = _BODY_FONT
                ws.cell(row=row, column=2, value=conflict.get("doc_target", "—")).font = _BODY_FONT
                ws.cell(row=row, column=3, value=conflict.get("conflict_type", conflict.get("type", "—"))).font = _BODY_FONT
                ws.cell(row=row, column=4, value=conflict.get("severity", "—")).font = _BOLD_FONT
                ws.cell(row=row, column=4).alignment = _CENTER
                ws.cell(row=row, column=5, value=conflict.get("description", "—")).font = _BODY_FONT
                ws.cell(row=row, column=5).alignment = _WRAP
                _style_data_row(ws, row, len(headers), alt=i % 2 == 1)
                row += 1

    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_analysis_excel(db: Session, project_id: str) -> bytes:
    """Génère un classeur Excel multi-onglets avec tous les résultats d'analyse.

    Onglets:
    1. Synthèse — infos projet, Go/No-Go, checklist stats
    2. Checklist — tous les items avec filtres
    3. Scoring — simulation de notation + graphique
    4. Trésorerie — cashflow mensuel + graphique
    5. DPGF Pricing — benchmark prix
    6. Risques — CCAP + conflits

    Returns:
        bytes du fichier .xlsx
    """
    data = fetch_export_data(db, project_id)

    wb = Workbook()

    _add_synthese_sheet(wb, data)
    _add_checklist_sheet(wb, data)
    _add_scoring_sheet(wb, data)
    _add_cashflow_sheet(wb, data)
    _add_dpgf_pricing_sheet(wb, data)
    _add_risks_sheet(wb, data)

    # Remove empty default sheet if we have others
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
